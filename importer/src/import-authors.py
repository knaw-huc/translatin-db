#!/usr/bin/env python3

from icecream import ic
from openpyxl import load_workbook

import psycopg2.extras
from psycopg2.extensions import AsIs
from psycopg2.extras import execute_values

import uuid

from mapping.authors import AUTHOR_ORIGIN, AUTHOR_STD_NAME, AUTHOR_TYPE, \
    AUTHOR_BIRTH_EARLIEST, AUTHOR_BIRTH_LATEST, AUTHOR_BIRTH_PLACE, \
    AUTHOR_DEATH_EARLIEST, AUTHOR_DEATH_LATEST, AUTHOR_DEATH_PLACE, \
    AUTHOR_ALT_NAME_FROM, AUTHOR_ALT_NAME_UPTO, AUTHOR_OCCUPATION, \
    AUTHOR_VIAF_FROM, AUTHOR_VIAF_UPTO, \
    AUTHOR_RELIGION, AUTHOR_IMAGE, AUTHOR_WIKIDATA

wb = load_workbook("/Users/jong/prj/translatin/download/TransLatin_Authors.xlsx")
ic(wb.sheetnames)
sheet = wb['Authors']


def show_titles():
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for i in range(0, len(row)):
            print(f"col[{i}]={row[i]}")


def collect_places(cursor):
    places = set()  # remove possible duplicates by collecting in a set
    for row in sheet.iter_rows(min_row=2, values_only=True):  # skip title row
        if row[AUTHOR_BIRTH_PLACE]:
            places.add(row[AUTHOR_BIRTH_PLACE])
        if row[AUTHOR_DEATH_PLACE]:
            places.add(row[AUTHOR_DEATH_PLACE])

    stmt = "INSERT INTO places (name) VALUES %s"
    data = [(place,) for place in places]  # convert set to list of tuples
    execute_values(cursor, stmt, data)


def create_authors(cursor):
    from datetime import date
    for row in sheet.iter_rows(min_row=2, values_only=True):  # skip title row
        try:
            author = {
                'origin': ic(row[AUTHOR_ORIGIN]),
                'name': row[AUTHOR_STD_NAME],
                'type': row[AUTHOR_TYPE]
            }

            # birth related data
            if row[AUTHOR_BIRTH_EARLIEST]:
                author['birth_earliest'] = date.fromisoformat(row[AUTHOR_BIRTH_EARLIEST])
            if row[AUTHOR_BIRTH_LATEST]:
                author['birth_latest'] = date.fromisoformat(row[AUTHOR_BIRTH_LATEST])
            if row[AUTHOR_BIRTH_PLACE]:
                author['_birth_place'] = row[AUTHOR_BIRTH_PLACE]

            # death related data
            if row[AUTHOR_DEATH_EARLIEST]:
                author['death_earliest'] = date.fromisoformat(row[AUTHOR_DEATH_EARLIEST])
            if row[AUTHOR_DEATH_LATEST]:
                author['death_latest'] = date.fromisoformat(row[AUTHOR_DEATH_LATEST])
            if row[AUTHOR_BIRTH_PLACE]:
                author['_death_place'] = row[AUTHOR_DEATH_PLACE]

            if row[AUTHOR_OCCUPATION]:
                author['occupation'] = row[AUTHOR_OCCUPATION]  # split on ';' and convert to 1:n relationship?

            if row[AUTHOR_RELIGION]:
                author['religion'] = row[AUTHOR_RELIGION]  # convert to enum type?

            if row[AUTHOR_IMAGE]:
                author['image'] = row[AUTHOR_IMAGE]

            if row[AUTHOR_WIKIDATA]:
                author['wikidata'] = row[AUTHOR_WIKIDATA]

            # 1:n relationship with alternative literal names
            names = [row[i] for i in range(AUTHOR_ALT_NAME_FROM, AUTHOR_ALT_NAME_UPTO) if row[i]]
            author['_names'] = fix_duplicates(row[AUTHOR_ORIGIN], names)

            # 1:n relationship with VIAF links
            viaf = [row[i] for i in range(AUTHOR_VIAF_FROM, AUTHOR_VIAF_UPTO) if row[i]]
            author['_viaf_links'] = fix_duplicates(row[AUTHOR_ORIGIN], viaf)

            create_author(cursor, author)
        except ValueError as err:
            ic(row[AUTHOR_ORIGIN], "illegal date", err)


def fix_duplicates(origin, some_list):
    normalised = list(dict.fromkeys(some_list))  # as of Python 3.7 also maintains original insertion order
    if normalised != some_list:
        ic('FIXING DUPLICATE: ', origin, some_list)
    return normalised


def create_author(cursor, author):
    # skip orgs for now
    if author['type'] == 'Organization':
        return

    if 'id' not in author:
        author['id'] = uuid.uuid4()

    stmt = 'INSERT INTO authors (%s) VALUES %s RETURNING id'
    columns = [col for col in author.keys() if not col.startswith('_')]
    values = [author[col] for col in columns]
    data = (AsIs(','.join(columns)), tuple(values))
    cursor.execute(stmt, data)
    author_id = cursor.fetchone()[0]

    if '_birth_place' in author:
        stmt = 'UPDATE authors SET birth_place = (SELECT id FROM places WHERE name = %s) WHERE id = %s'
        data = (author['_birth_place'], author_id)
        # ic(cursor.mogrify(stmt, data))
        cursor.execute(stmt, data)

    stmt = 'INSERT INTO author_names (author_id, name) VALUES %s'
    data = [(author['id'], name) for name in author['_names']]
    execute_values(cursor, stmt, data)

    stmt = 'INSERT INTO author_viaf_links (author_id, viaf) VALUES %s'
    data = [(author['id'], viaf) for viaf in author['_viaf_links']]
    execute_values(cursor, stmt, data)


conn = None
try:
    print("Connecting to translatin database...")
    psycopg2.extras.register_uuid()
    conn = psycopg2.connect(host="localhost",
                            database="translatin",
                            user="translatin",
                            password="translatin")
    with conn.cursor() as curs:
        curs.execute("select version()")
        version = curs.fetchone()
        ic(version)
        # save_manifestation(curs, mock())
        collect_places(curs)
        conn.commit()
        create_authors(curs)
        conn.commit()
except (Exception, psycopg2.DatabaseError) as error:
    print(error)
finally:
    if conn is not None:
        conn.close()
        print('Database connection closed.')

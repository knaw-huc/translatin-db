#!/usr/bin/env python3

import configparser
from icecream import ic
from openpyxl import load_workbook
from util import show_titles

parser = configparser.ConfigParser(interpolation=configparser.ExtendedInterpolation())
parser.read('config.ini')
conf = parser['manifestations']
wb = load_workbook(ic(conf['path']))
ic(wb.sheetnames)
sheet = wb[conf['name']]

show_titles(sheet)
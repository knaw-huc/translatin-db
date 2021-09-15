#!/usr/bin/env python3

from icecream import ic
from openpyxl import load_workbook

import psycopg2.extras
from psycopg2.extensions import AsIs
from psycopg2.extras import execute_values

import uuid

from mapping import MF_ORIGIN, MF_CENETON_FROM, MF_CENETON_UPTO, MF_EARLIEST, MF_LATEST, MF_FINGERPRINT, \
    MF_LANG_FROM, MF_LANG_UPTO, MF_TITLE_FROM, MF_TITLE_UPTO, MF_FORM, MF_FORM_TYPE, MF_GENRE, MF_SUBGENRE, \
    MF_CHARACTERS, MF_REMARKS, MF_LITERATURE, MF_HAS_TRANSCRIPTION

wb = load_workbook("/Users/jong/prj/translatin/download/TransLatin_Manifestations.xlsx")
ic(wb.sheetnames)

sheet = wb['Blad1']
pers = ic(sheet['CH39'].value)
ic(pers.split('_x000B_'))  # vertical tab \u000B encoded


def create_manifestations(cursor):
    # iterate over all rows after title row
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # mandatory fields, usable 'as is'
        man = {
            'origin': ic(row[MF_ORIGIN]),
            'earliest': row[MF_EARLIEST],
            'latest': row[MF_LATEST],
            'form_type': row[MF_FORM_TYPE]
        }

        # mandatory field requiring a massage: 'has_transcription'. Defaults to False for empty cells,
        # has 'YES' for True and '?' for to-be-determined which is NULL in db
        if row[MF_HAS_TRANSCRIPTION]:
            man['has_transcription'] = True if row[MF_HAS_TRANSCRIPTION] == 'YES' else None
        else:
            man['has_transcription'] = False

        # optional fields
        if row[MF_FORM]:
            man['form'] = fix_form(row[MF_FORM])
        if row[MF_FINGERPRINT]:
            man['fingerprint'] = row[MF_FINGERPRINT]
        if row[MF_GENRE]:
            man['genre'] = row[MF_GENRE]
        if row[MF_SUBGENRE]:
            man['subgenre'] = row[MF_SUBGENRE]
        if row[MF_CHARACTERS]:
            man['characters'] = row[MF_CHARACTERS]
        if row[MF_REMARKS]:
            man['remarks'] = row[MF_REMARKS]
        if row[MF_LITERATURE]:
            man['literature'] = row[MF_LITERATURE]

        # 1:n relationship with Ceneton identifiers
        ceneton_ids = [row[i] for i in range(MF_CENETON_FROM, MF_CENETON_UPTO) if row[i]]
        man['_ceneton'] = fix_duplicates(ceneton_ids)

        # 1:n relationship with titles
        titles = [row[i] for i in range(MF_TITLE_FROM, MF_TITLE_UPTO) if row[i]]
        man['_titles'] = fix_duplicates(titles)

        # 1:n relationship with (language, certainty) aka 'Language' pairs
        languages = [(fix_language(row[i]), row[i + 1]) for i in range(MF_LANG_FROM, MF_LANG_UPTO, 2) if row[i]]
        man['_languages'] = fix_duplicates(languages)

        create_manifestation(cursor, man)


def fix_duplicates(some_list):
    return list(dict.fromkeys(some_list))  # as of Python 3.7 also maintains original insertion order


def fix_form(form):
    if form == 'Synopsis (gedrukt)':
        return 'Synopsis (printed)'
    return form


def fix_language(lang):
    if lang == 'Nederlands':
        return 'Dutch'
    return lang


def create_manifestation(cursor, man):
    if 'id' not in man:
        man['id'] = uuid.uuid4()

    stmt = 'INSERT INTO manifestations (%s) VALUES %s'
    columns = [col for col in man.keys() if not col.startswith('_')]
    values = [man[col] for col in columns]
    data = (AsIs(','.join(columns)), tuple(values))
    cursor.execute(stmt, data)

    stmt = 'INSERT INTO manifestation_ceneton (manifestation_id, ceneton_id) VALUES %s'
    data = [(man['id'], cid) for cid in man['_ceneton']]
    execute_values(cursor, stmt, data)

    stmt = 'INSERT INTO manifestation_titles (manifestation_id, title) VALUES %s'
    data = [(man['id'], title) for title in man['_titles']]
    execute_values(cursor, stmt, data)

    stmt = 'INSERT INTO manifestation_languages (manifestation_id, language, certainty) VALUES %s'
    data = [(man['id'], lang, cert) for lang, cert in man['_languages']]
    execute_values(cursor, stmt, data)


def show_titles():
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for i in range(0, len(row)):
            print(f"col[{i}]={row[i]}")


conn = None
try:
    print("Connecting to translatin database...")
    psycopg2.extras.register_uuid()
    conn = psycopg2.connect(host="localhost",
                            database="translatin",
                            user="translatin",
                            password="translatin")
    with conn.cursor() as curs:
        curs.execute("select version()")
        version = curs.fetchone()
        ic(version)
        # save_manifestation(curs, mock())
        create_manifestations(curs)
        conn.commit()
except (Exception, psycopg2.DatabaseError) as error:
    print(error)
finally:
    if conn is not None:
        conn.close()
        print('Database connection closed.')

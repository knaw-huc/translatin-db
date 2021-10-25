#!/usr/bin/env python3

import configparser
from icecream import ic
from openpyxl import load_workbook

import psycopg2.extras
from psycopg2.extensions import AsIs
from psycopg2.extras import execute_values

from util import fix_duplicates

import uuid

from mapping.manifestations import MF_ORIGIN, MF_CENETON_FROM, MF_CENETON_UPTO, MF_EARLIEST, MF_LATEST, \
    MF_AUTHOR_FROM, MF_AUTHOR_UPTO, MF_FINGERPRINT, MF_TITLE_FROM, MF_LANG_FROM, MF_CERT_FROM, MF_TITLE_UPTO, \
    MF_FORM, MF_FORM_TYPE, MF_PUBLISHER_FROM, MF_PUBLISHER_UPTO, MF_GENRE, MF_SUBGENRE, MF_CHARACTERS, MF_REMARKS, \
    MF_LITERATURE, MF_HAS_DRAMAWEB_TRANSCRIPTION, MF_EXTERNAL_SCAN_URL, MF_CENETON_SCAN_URL, \
    MF_CENETON_TRANSCRIPTION_URL, MF_PUBLISH_PLACE_FROM, MF_PUBLISH_PLACE_UPTO

parser = configparser.ConfigParser(interpolation=configparser.ExtendedInterpolation())
parser.read('config.ini')
conf = parser['manifestations']
wb = load_workbook(ic(conf['path']))
ic(wb.sheetnames)
sheet = wb[conf['name']]


def collect_places(cursor):
    places = set()  # remove possible duplicates by collecting into a set

    for row in sheet.iter_rows(min_row=2, values_only=True):  # skip title row
        for i in range(MF_PUBLISH_PLACE_FROM, MF_PUBLISH_PLACE_UPTO):
            if row[i]:
                places.add(row[i])

    stmt = "INSERT INTO places (name) VALUES %s ON CONFLICT (name) DO NOTHING"
    data = [(place,) for place in places]  # convert set to list of tuples
    execute_values(cursor, stmt, data)


def create_manifestations(cursor):
    # iterate over all rows after title row
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # mandatory fields, usable 'as is'
        man = {
            'origin': ic(row[MF_ORIGIN]),
            'earliest': row[MF_EARLIEST],
            'latest': row[MF_LATEST],
            'form_type': row[MF_FORM_TYPE]
        }

        # mandatory field requiring a massage: 'has_transcription'. Defaults to False for empty cells,
        # has 'YES' for True and '?' for to-be-determined which is NULL in db
        if row[MF_HAS_DRAMAWEB_TRANSCRIPTION]:
            man['has_dramaweb_transcription'] = True if row[MF_HAS_DRAMAWEB_TRANSCRIPTION] == 'YES' else None
        else:
            man['has_dramaweb_transcription'] = False

        # for now we assume that if there is a transcription, we also have the scan
        man['has_dramaweb_scan'] = man['has_dramaweb_transcription']

        # optional fields
        if row[MF_FORM]:
            man['form'] = row[MF_FORM]
        if row[MF_FINGERPRINT]:
            man['fingerprint'] = row[MF_FINGERPRINT]
        if row[MF_GENRE]:
            man['genre'] = row[MF_GENRE]
        if row[MF_SUBGENRE]:
            man['subgenre'] = row[MF_SUBGENRE]
        if row[MF_CHARACTERS]:
            # compound cell that has 'personages' separated by '_x000B_'
            personages = row[MF_CHARACTERS].split('_x000B_')
            # sometimes starts with '(Geen afzonderlijke lijst van personages)' which we then remove
            if personages[0] == '(Geen afzonderlijke lijst van personages)':
                personages.pop(0)
            man['_personages'] = fix_duplicates(row[MF_ORIGIN], personages)
        if row[MF_REMARKS]:
            man['remarks'] = row[MF_REMARKS]
        if row[MF_LITERATURE]:
            man['literature'] = row[MF_LITERATURE]
        if row[MF_CENETON_SCAN_URL]:
            val = row[MF_CENETON_SCAN_URL]
            if len(val) < 8:
                ic(row[MF_ORIGIN], 'SUSPICIOUSLY SHORT CENETON SCAN REF', val)
            if '_x000B_' in val:
                ic('VERTICAL SPACE (\\v, 0x0b) in:', val)
                val = val.replace('_x000B_', '')
            man['ceneton_scan'] = f'https://www.let.leidenuniv.nl/Dutch/Ceneton/Facsimiles/{val}'
        if row[MF_CENETON_TRANSCRIPTION_URL]:
            val = row[MF_CENETON_TRANSCRIPTION_URL]
            if len(val) < 8:
                ic(row[MF_ORIGIN], 'SUSPICIOUSLY SHORT CENETON TRANSCRIPTION REF', val)
            if '_x000B_' in val:
                ic(row[MF_ORIGIN], 'VERTICAL SPACE (\\v, 0x0b) in:', val)
                val = val.replace('_x000B_', '')
            man['ceneton_transcription'] = f'https://www.let.leidenuniv.nl/Dutch/Ceneton/{val}.html'
        if row[MF_EXTERNAL_SCAN_URL]:
            man['external_scan'] = row[MF_EXTERNAL_SCAN_URL]

        # not in excel sheet, assume false. Can manually be set to True. Perhaps make this 'text' to hold URLs?
        man['external_transcription'] = False

        # 1:n relationship with Ceneton identifiers
        ceneton_ids = [row[i] for i in range(MF_CENETON_FROM, MF_CENETON_UPTO) if row[i]]
        man['_ceneton'] = fix_duplicates(row[MF_ORIGIN], ceneton_ids)

        # 1:n relationship with titles which are {certain,uncertain} to be of a specific language
        titles = [(row[MF_TITLE_FROM + i], row[MF_LANG_FROM + (2 * i)], row[MF_CERT_FROM + (2 * i)])
                  for i in range(0, MF_TITLE_UPTO - MF_TITLE_FROM) if row[MF_TITLE_FROM + i]]
        man['_titles'] = titles

        # 1:n relationship with authors. As these are linked by name in Excel, authors MUST be imported first!
        # Also keep track of author type ('Person', 'Organization') as this must match during lookup in db.
        authors = [(row[i], row[i + 1]) for i in range(MF_AUTHOR_FROM, MF_AUTHOR_UPTO, 2) if row[i]]
        if ('Anonymous', 'Unknown') in authors:
            man['is_anonymous'] = True
        else:
            man['is_anonymous'] = False
            man['_authors'] = fix_authors(cursor, row[MF_ORIGIN], authors)

        # 1:n relationship with publishers/printers.
        # These are also linked by name in Excel, so publishers MUST be imported first.
        place_offset = MF_PUBLISH_PLACE_FROM - MF_PUBLISHER_FROM
        publishers = [(row[i], row[i + place_offset]) for i in range(MF_PUBLISHER_FROM, MF_PUBLISHER_UPTO) if row[i]]
        man['_publishers'] = fix_publishers(cursor, row[MF_ORIGIN], publishers)

        create_manifestation(cursor, man)


def fix_publishers(cursor, origin, publishers):
    unique_names = set()
    fixed_publishers = list()

    for (publisher_name, place_name) in publishers:
        if not place_name:
            ic('MISSING PUBLICATION PLACE', origin, publisher_name)

        if publisher_name in unique_names:
            ic('DUPLICATE PUBLISHER', origin, publisher_name)
            continue
        else:
            unique_names.add(publisher_name)

        cursor.execute('SELECT EXISTS(SELECT 1 FROM publishers WHERE name = %s)', (publisher_name,))
        if not cursor.fetchone()[0]:
            ic('UNKNOWN PUBLISHER NAME', origin, publisher_name)
            continue

        fixed_publishers.append((publisher_name, place_name))

    return fixed_publishers


def fix_authors(cursor, origin, authors):
    unique_names = set()
    fixed_authors = list()

    for (author_name, author_type) in authors:
        if author_name in unique_names:
            ic('DUPLICATE AUTHOR NAME', origin, author_name)
            continue
        else:
            unique_names.add(author_name)

        cursor.execute('SELECT EXISTS(SELECT 1 FROM authors WHERE name = %s AND type = %s)',
                       (author_name, author_type))
        if not cursor.fetchone()[0]:
            ic('UNKNOWN AUTHOR+TYPE', origin, author_name, author_type)
            cursor.execute('SELECT EXISTS(SELECT 1 FROM authors WHERE name = %s)', (author_name,))
            if cursor.fetchone()[0]:
                ic('AUTHOR EXISTS WITH DIFFERENT TYPE', origin, author_name)
            continue

        fixed_authors.append((author_name, author_type))

    return fixed_authors


def create_manifestation(cursor, man):
    if 'id' not in man:
        man['id'] = uuid.uuid4()

    stmt = 'INSERT INTO manifestations (%s) VALUES %s'
    columns = [col for col in man.keys() if not col.startswith('_')]
    values = [man[col] for col in columns]
    data = (AsIs(','.join(columns)), tuple(values))
    cursor.execute(stmt, data)

    stmt = 'INSERT INTO manifestation_ceneton (manifestation_id, ceneton_id) VALUES %s'
    data = [(man['id'], cid) for cid in man['_ceneton']]
    execute_values(cursor, stmt, data)

    stmt = 'INSERT INTO manifestation_titles (manifestation_id, title, language, certainty) VALUES %s'
    data = [(man['id'], title, lang, cert) for title, lang, cert in man['_titles']]
    execute_values(cursor, stmt, data)

    if '_personages' in man:
        stmt = 'INSERT INTO manifestation_personages (manifestation_id, name) VALUES %s'
        data = [(man['id'], name) for name in man['_personages'] if name]
        execute_values(cursor, stmt, data)

    if '_authors' in man:
        for (author_name, author_type) in man['_authors']:
            stmt = '''
            INSERT INTO authors_manifestations (author_id, manifestation_id) VALUES (
                (SELECT id FROM authors WHERE name = %s AND type = %s),
                %s)
            '''
            data = (author_name, author_type, man['id'])
            cursor.execute(stmt, data)

    if '_publishers' in man:
        for (publisher_name, place_name) in man['_publishers']:
            if place_name:
                stmt = '''
                INSERT INTO manifestations_publishers (manifestation_id, publisher_id, place_id) VALUES (
                    %s,
                    (SELECT id FROM publishers WHERE name = %s),
                    (SELECT id FROM places WHERE name = %s))
                '''
                data = (man['id'], publisher_name, place_name)
            else:
                stmt = '''
                INSERT INTO manifestations_publishers (manifestation_id, publisher_id) VALUES (
                    %s,
                    (SELECT id FROM publishers WHERE name = %s))
                '''
                data = (man['id'], publisher_name)
            cursor.execute(stmt, data)


conn = None
try:
    print("Connecting to translatin database...")
    psycopg2.extras.register_uuid()
    conn = psycopg2.connect(**parser['db'])
    with conn.cursor() as curs:
        curs.execute("select version()")
        version = curs.fetchone()
        ic(version)
        collect_places(curs)
        create_manifestations(curs)
        conn.commit()
except (Exception, psycopg2.DatabaseError) as error:
    print(error)
finally:
    if conn is not None:
        conn.close()
        print('Database connection closed.')

#!/usr/bin/env python3

import configparser
import os
import uuid

import psycopg2.extras
from icecream import ic
from openpyxl import load_workbook
from psycopg2.extensions import AsIs
from psycopg2.extras import execute_values

from mapping.publishers import PP_STD_NAME, PP_WIDOW_HEIRS, \
    PP_FIRST_NAME, PP_PATRONYM, PP_PREFIX, PP_SURNAME, PP_ADDITION, \
    PP_ALT_NAMES_FROM, PP_ALT_NAMES_UPTO, PP_CERL_LINK
from util import fix_duplicates, sha1sum

parser = configparser.ConfigParser(interpolation=configparser.ExtendedInterpolation())
parser.read('config.ini')
conf = parser['publishers']
path = ic(conf['path'])
prov_name = ic(os.path.basename(path))
prov_size = ic(os.path.getsize(path))
prov_sha1 = ic(sha1sum(path))
wb = load_workbook(path)
ic(wb.sheetnames)
sheet = wb[conf['name']]

# show_titles(sheet)

conn = None
unique_names = set()


def create_publishers(cursor):
    for row in sheet.iter_rows(min_row=2, values_only=True):  # skip title row
        name = row[PP_STD_NAME]
        ic(name)

        # avoid postgres zapping us on duplicate names
        if name in unique_names:
            ic('DUPLICATE NAME', name)
            continue
        unique_names.add(name)

        publisher = {
            'name': name,
            'surname': row[PP_SURNAME]
        }

        # Name related data
        if row[PP_WIDOW_HEIRS]:
            publisher['widow_heirs'] = row[PP_WIDOW_HEIRS]
        if row[PP_FIRST_NAME]:
            publisher['first_name'] = row[PP_FIRST_NAME]
        if row[PP_PATRONYM]:
            publisher['patronym'] = row[PP_PATRONYM]
        if row[PP_PREFIX]:
            publisher['prefix'] = row[PP_PREFIX]
        if row[PP_ADDITION]:
            publisher['addition'] = row[PP_ADDITION]
        if row[PP_CERL_LINK]:
            publisher['cerl_link'] = row[PP_CERL_LINK]

        # 1:n relationship with alternative literal names
        names = [row[i] for i in range(PP_ALT_NAMES_FROM, PP_ALT_NAMES_UPTO) if row[i]]
        publisher['_names'] = fix_duplicates(name, names)

        create_publisher(cursor, publisher)


def create_publisher(cursor, publisher):
    if 'id' not in publisher:
        publisher['id'] = uuid.uuid4()

    stmt = 'INSERT INTO publishers (%s) VALUES %s'
    columns = [col for col in publisher.keys() if not col.startswith('_')]
    values = [publisher[col] for col in columns]
    data = (AsIs(','.join(columns)), tuple(values))
    cursor.execute(stmt, data)

    stmt = 'INSERT INTO publisher_names (publisher_id, name) VALUES %s'
    data = [(publisher['id'], name) for name in publisher['_names']]
    execute_values(cursor, stmt, data)


def update_provenance(cursor):
    stmt = 'UPDATE _provenance SET size = %s, sha1sum = %s, imported_at = NOW() WHERE name = %s'
    data = [(prov_size, prov_sha1, prov_name)]
    execute_values(cursor, stmt, data)


try:
    print("Connecting to translatin database...")
    psycopg2.extras.register_uuid()
    conn = psycopg2.connect(**parser['db'])
    with conn.cursor() as curs:
        curs.execute("select version()")
        version = curs.fetchone()
        ic(version)
        create_publishers(curs)
        conn.commit()
        update_provenance(curs)
        conn.commit()
except (Exception, psycopg2.DatabaseError) as error:
    print(error)
finally:
    if conn is not None:
        conn.close()
        print('Database connection closed.')
